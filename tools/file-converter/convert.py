"""
File Converter Tool — Convert between file formats for AI agents.

Accepts file content via HTTP POST and converts between formats:
CSV, TSV, JSON, PDF, HTML, Markdown, XLSX, XML.
Returns structured data agents can parse and reason about.
"""

import os
import io
import csv
import json
import base64
import xml.etree.ElementTree as ET
from html.parser import HTMLParser
from http.server import HTTPServer, BaseHTTPRequestHandler

DEFAULT_PORT = 8003

# Supported conversion paths: (from, to) -> handler function
CONVERSIONS = {}


def conversion(from_fmt, to_fmt):
    """Decorator to register a conversion function."""
    def decorator(func):
        CONVERSIONS[(from_fmt, to_fmt)] = func
        return func
    return decorator


# ---------------------------------------------------------------------------
# Text extractors
# ---------------------------------------------------------------------------

class _HTMLTextExtractor(HTMLParser):
    """Strip HTML tags and return plain text."""
    def __init__(self):
        super().__init__()
        self._parts = []
        self._skip = False

    def handle_starttag(self, tag, attrs):
        if tag in ("script", "style"):
            self._skip = True

    def handle_endtag(self, tag):
        if tag in ("script", "style"):
            self._skip = False
        if tag in ("p", "div", "br", "li", "h1", "h2", "h3", "h4", "h5", "h6", "tr"):
            self._parts.append("\n")

    def handle_data(self, data):
        if not self._skip:
            self._parts.append(data)

    def get_text(self):
        return "".join(self._parts).strip()


def _strip_markdown(text: str) -> str:
    """Lightweight Markdown-to-plain-text conversion (no dependencies)."""
    import re
    text = re.sub(r"^#{1,6}\s+", "", text, flags=re.MULTILINE)
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    text = re.sub(r"\*(.+?)\*", r"\1", text)
    text = re.sub(r"__(.+?)__", r"\1", text)
    text = re.sub(r"_(.+?)_", r"\1", text)
    text = re.sub(r"~~(.+?)~~", r"\1", text)
    text = re.sub(r"`{1,3}(.+?)`{1,3}", r"\1", text, flags=re.DOTALL)
    text = re.sub(r"!?\[([^\]]*)\]\([^\)]+\)", r"\1", text)
    text = re.sub(r"^>\s?", "", text, flags=re.MULTILINE)
    text = re.sub(r"^[-*+]\s+", "- ", text, flags=re.MULTILINE)
    text = re.sub(r"^\d+\.\s+", "", text, flags=re.MULTILINE)
    text = re.sub(r"^---+$", "", text, flags=re.MULTILINE)
    return text.strip()


# ---------------------------------------------------------------------------
# CSV / TSV conversions
# ---------------------------------------------------------------------------

@conversion("csv", "json")
def csv_to_json(content: str, **_) -> dict:
    reader = csv.DictReader(io.StringIO(content))
    rows = list(reader)
    return {"data": rows, "rows": len(rows), "columns": reader.fieldnames or []}


@conversion("tsv", "json")
def tsv_to_json(content: str, **_) -> dict:
    reader = csv.DictReader(io.StringIO(content), delimiter="\t")
    rows = list(reader)
    return {"data": rows, "rows": len(rows), "columns": reader.fieldnames or []}


@conversion("json", "csv")
def json_to_csv(content: str, **_) -> dict:
    data = json.loads(content)
    if isinstance(data, dict) and "data" in data:
        data = data["data"]
    if not isinstance(data, list) or len(data) == 0:
        return {"error": "JSON must be an array of objects (or {\"data\": [...]})"}

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=list(data[0].keys()))
    writer.writeheader()
    writer.writerows(data)
    return {"data": output.getvalue(), "rows": len(data), "format": "text/csv"}


@conversion("csv", "tsv")
def csv_to_tsv(content: str, **_) -> dict:
    reader = csv.reader(io.StringIO(content))
    output = io.StringIO()
    writer = csv.writer(output, delimiter="\t")
    rows = 0
    for row in reader:
        writer.writerow(row)
        rows += 1
    return {"data": output.getvalue(), "rows": rows, "format": "text/tsv"}


# ---------------------------------------------------------------------------
# HTML conversions
# ---------------------------------------------------------------------------

@conversion("html", "text")
def html_to_text(content: str, **_) -> dict:
    extractor = _HTMLTextExtractor()
    extractor.feed(content)
    text = extractor.get_text()
    return {"data": text, "characters": len(text)}


@conversion("html", "json")
def html_to_json(content: str, **_) -> dict:
    """Extract tables from HTML into JSON arrays."""
    import re
    tables = []
    table_pattern = re.compile(r"<table[^>]*>(.*?)</table>", re.DOTALL | re.IGNORECASE)
    row_pattern = re.compile(r"<tr[^>]*>(.*?)</tr>", re.DOTALL | re.IGNORECASE)
    cell_pattern = re.compile(r"<t[hd][^>]*>(.*?)</t[hd]>", re.DOTALL | re.IGNORECASE)
    tag_strip = re.compile(r"<[^>]+>")

    for table_match in table_pattern.finditer(content):
        rows = []
        headers = []
        for i, row_match in enumerate(row_pattern.finditer(table_match.group(1))):
            cells = [tag_strip.sub("", c).strip() for c in cell_pattern.findall(row_match.group(1))]
            if i == 0:
                headers = cells
            else:
                rows.append(dict(zip(headers, cells)) if headers else cells)
        tables.append({"headers": headers, "rows": rows, "row_count": len(rows)})

    if not tables:
        extractor = _HTMLTextExtractor()
        extractor.feed(content)
        return {"data": extractor.get_text(), "tables": [], "note": "No tables found, returning plain text."}

    return {"tables": tables, "table_count": len(tables)}


# ---------------------------------------------------------------------------
# Markdown conversions
# ---------------------------------------------------------------------------

@conversion("markdown", "text")
def markdown_to_text(content: str, **_) -> dict:
    text = _strip_markdown(content)
    return {"data": text, "characters": len(text)}


@conversion("markdown", "html")
def markdown_to_html(content: str, **_) -> dict:
    """Basic Markdown-to-HTML (covers common patterns, no external deps)."""
    import re
    html = content

    # Code blocks (``` ... ```)
    html = re.sub(r"```(\w*)\n(.*?)```", r"<pre><code>\2</code></pre>", html, flags=re.DOTALL)
    # Inline code
    html = re.sub(r"`([^`]+)`", r"<code>\1</code>", html)
    # Headers
    for i in range(6, 0, -1):
        html = re.sub(rf"^{'#' * i}\s+(.+)$", rf"<h{i}>\1</h{i}>", html, flags=re.MULTILINE)
    # Bold / italic
    html = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", html)
    html = re.sub(r"\*(.+?)\*", r"<em>\1</em>", html)
    # Links
    html = re.sub(r"\[([^\]]+)\]\(([^\)]+)\)", r'<a href="\2">\1</a>', html)
    # Line breaks
    html = re.sub(r"\n\n", "</p><p>", html)
    html = f"<p>{html}</p>"

    return {"data": html, "format": "text/html"}


# ---------------------------------------------------------------------------
# XML conversions
# ---------------------------------------------------------------------------

def _xml_to_dict(element):
    """Recursively convert an XML element to a dict."""
    result = {}
    if element.attrib:
        result["@attributes"] = dict(element.attrib)
    children = list(element)
    if children:
        child_dict = {}
        for child in children:
            key = child.tag
            value = _xml_to_dict(child)
            if key in child_dict:
                if not isinstance(child_dict[key], list):
                    child_dict[key] = [child_dict[key]]
                child_dict[key].append(value)
            else:
                child_dict[key] = value
        result.update(child_dict)
    elif element.text and element.text.strip():
        if result:
            result["#text"] = element.text.strip()
        else:
            return element.text.strip()
    return result


@conversion("xml", "json")
def xml_to_json(content: str, **_) -> dict:
    try:
        root = ET.fromstring(content)
    except ET.ParseError as e:
        return {"error": f"Invalid XML: {e}"}
    return {"data": {root.tag: _xml_to_dict(root)}}


@conversion("json", "xml")
def json_to_xml(content: str, **_) -> dict:
    data = json.loads(content)

    def _dict_to_xml(d, parent):
        if isinstance(d, dict):
            for key, val in d.items():
                if key.startswith("@"):
                    continue
                if isinstance(val, list):
                    for item in val:
                        child = ET.SubElement(parent, key)
                        _dict_to_xml(item, child)
                else:
                    child = ET.SubElement(parent, key)
                    _dict_to_xml(val, child)
        else:
            parent.text = str(d)

    if isinstance(data, dict) and len(data) == 1:
        root_key = list(data.keys())[0]
        root = ET.Element(root_key)
        _dict_to_xml(data[root_key], root)
    else:
        root = ET.Element("root")
        _dict_to_xml(data, root)

    xml_str = ET.tostring(root, encoding="unicode", xml_declaration=True)
    return {"data": xml_str, "format": "text/xml"}


# ---------------------------------------------------------------------------
# PDF conversion (requires PyPDF2 — optional dependency)
# ---------------------------------------------------------------------------

@conversion("pdf", "text")
def pdf_to_text(content: str, **_) -> dict:
    try:
        import PyPDF2
    except ImportError:
        return {"error": "PyPDF2 is not installed. Run: pip install PyPDF2"}

    try:
        pdf_bytes = base64.b64decode(content)
    except Exception:
        return {"error": "PDF content must be base64-encoded."}

    reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        pages.append({"page": i + 1, "text": text})

    full_text = "\n\n".join(p["text"] for p in pages)
    return {"data": full_text, "pages": pages, "page_count": len(pages), "characters": len(full_text)}


@conversion("pdf", "json")
def pdf_to_json(content: str, **_) -> dict:
    result = pdf_to_text(content)
    if "error" in result:
        return result
    return {"pages": result["pages"], "page_count": result["page_count"]}


# ---------------------------------------------------------------------------
# XLSX conversion (requires openpyxl — optional dependency)
# ---------------------------------------------------------------------------

@conversion("xlsx", "json")
def xlsx_to_json(content: str, **_) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl is not installed. Run: pip install openpyxl"}

    try:
        xlsx_bytes = base64.b64decode(content)
    except Exception:
        return {"error": "XLSX content must be base64-encoded."}

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            sheets.append({"sheet": name, "headers": [], "data": [], "rows": 0})
            continue
        headers = [str(h) if h is not None else f"column_{i}" for i, h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            data.append(dict(zip(headers, [v for v in row])))
        sheets.append({"sheet": name, "headers": headers, "data": data, "rows": len(data)})
    wb.close()

    return {"sheets": sheets, "sheet_count": len(sheets)}


@conversion("xlsx", "csv")
def xlsx_to_csv(content: str, **_) -> dict:
    result = xlsx_to_json(content)
    if "error" in result:
        return result

    outputs = []
    for sheet in result["sheets"]:
        output = io.StringIO()
        if sheet["data"]:
            writer = csv.DictWriter(output, fieldnames=sheet["headers"])
            writer.writeheader()
            writer.writerows(sheet["data"])
        outputs.append({"sheet": sheet["sheet"], "csv": output.getvalue(), "rows": sheet["rows"]})

    return {"sheets": outputs, "sheet_count": len(outputs), "format": "text/csv"}


# ---------------------------------------------------------------------------
# Utility: list supported conversions
# ---------------------------------------------------------------------------

def list_conversions() -> dict:
    paths = []
    for (from_fmt, to_fmt) in sorted(CONVERSIONS.keys()):
        paths.append({"from": from_fmt, "to": to_fmt})
    return {"conversions": paths, "count": len(paths)}


def convert(content: str, from_fmt: str, to_fmt: str) -> dict:
    """Run a conversion. Returns result dict or error dict."""
    from_fmt = from_fmt.lower().strip()
    to_fmt = to_fmt.lower().strip()

    # Normalize aliases
    aliases = {"md": "markdown", "htm": "html", "xls": "xlsx"}
    from_fmt = aliases.get(from_fmt, from_fmt)
    to_fmt = aliases.get(to_fmt, to_fmt)

    handler = CONVERSIONS.get((from_fmt, to_fmt))
    if not handler:
        return {
            "error": f"Unsupported conversion: {from_fmt} → {to_fmt}",
            "supported": list_conversions()["conversions"],
        }

    try:
        return handler(content)
    except json.JSONDecodeError as e:
        return {"error": f"Invalid JSON input: {e}"}
    except Exception as e:
        return {"error": f"Conversion failed: {type(e).__name__}: {e}"}


# ---------------------------------------------------------------------------
# HTTP Server
# ---------------------------------------------------------------------------

class ConvertHandler(BaseHTTPRequestHandler):
    """HTTP handler for file conversion."""

    def do_GET(self):
        import urllib.parse
        parsed = urllib.parse.urlparse(self.path)

        if parsed.path == "/conversions":
            self._respond(200, list_conversions())
            return

        self._respond(404, {
            "error": "Not found. Use POST /convert or GET /conversions.",
            "endpoints": {
                "POST /convert": "Convert file content (JSON body with 'content', 'from', 'to')",
                "GET /conversions": "List all supported conversion paths",
            }
        })

    def do_POST(self):
        import urllib.parse
        parsed = urllib.parse.urlparse(self.path)

        if parsed.path != "/convert":
            self._respond(404, {"error": "Not found. Use POST /convert"})
            return

        content_length = int(self.headers.get("Content-Length", 0))
        if content_length == 0:
            self._respond(400, {"error": "Empty request body"})
            return

        try:
            body = json.loads(self.rfile.read(content_length))
        except json.JSONDecodeError:
            self._respond(400, {"error": "Request body must be valid JSON"})
            return

        file_content = body.get("content")
        from_fmt = body.get("from", "")
        to_fmt = body.get("to", "")

        if not file_content:
            self._respond(400, {"error": "Missing required field: content"})
            return
        if not from_fmt:
            self._respond(400, {"error": "Missing required field: from"})
            return
        if not to_fmt:
            self._respond(400, {"error": "Missing required field: to"})
            return

        result = convert(file_content, from_fmt, to_fmt)
        status = 422 if "error" in result else 200
        self._respond(status, result)

    def _respond(self, status: int, body: dict):
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.end_headers()
        self.wfile.write(json.dumps(body, indent=2, default=str).encode())

    def log_message(self, fmt, *args):
        print(f"[file-converter] {args[0]}")


def main():
    port = int(os.environ.get("PORT", DEFAULT_PORT))
    server = HTTPServer(("0.0.0.0", port), ConvertHandler)
    print(f"File Converter running on http://localhost:{port}")
    print(f"  POST /convert  — convert file content")
    print(f"  GET  /conversions — list supported formats")
    print(f"\nSupported conversions:")
    for (f, t) in sorted(CONVERSIONS.keys()):
        print(f"  {f} → {t}")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nShutting down.")
        server.server_close()


if __name__ == "__main__":
    main()
